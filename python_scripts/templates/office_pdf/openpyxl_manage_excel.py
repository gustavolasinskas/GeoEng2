#!/usr/bin/env Python3

import openpyxl, os

wb = openpyxl.Workbook()
wb

wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Sheet')
sheet
sheet['A1'].value == None

sheet['A1'].value = 42
sheet['A2'].value = 'Hello'

os.chdir('/users/alex/desktop')
sheet2 = wb.create_sheet()
wb.get_sheet_names()
wb.save('excel_example_python.xlsx')

sheet2.title = "New Sheet"
wb.create_sheet(index=0, title="First Sheet Ever!")
wb.save('excel_example2_python.xlsx')

for i in range(1, 100):
    sheet2.cell(row = i, column = 2).value=i

wb.save('excel_example2_python.xlsx')
